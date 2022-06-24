import ConnectionDB as DB
import openpyxl

book = openpyxl.open("dataForimport.xlsx")
sheet_1 = book.worksheets[0]
sheet_2 = book.worksheets[1]
sheet_3 = book.worksheets[2]
sheet_4 = book.worksheets[3]
sheet_5 = book.worksheets[4]

def InsertProductTypes():
    for row in range(1, sheet_1.max_row):
        DB.cur.execute(f"INSERT INTO public.\"ProductTypes\" (\"Title\") VALUES('{sheet_1[row][1].value}')")
    DB.connection.commit()

def InsertMaterialTypes():
    for row in range(1, sheet_2.max_row):
        DB.cur.execute(f"INSERT INTO public.\"MaterialTypes\" (\"Title\") VALUES ('{sheet_2[row][1].value}')")
    DB.connection.commit()

def InsertMaterials():
    for row in range(1, sheet_3.max_row):
        DB.cur.execute(
            f"""INSERT INTO public."Materials" ("Title", "CountInPack", "Uint", "CountInStock", "MinCount", "Description", "Cost", "Image", "MaterialTypeId") VALUES ('{sheet_3[row][1].value}', {sheet_3[row][2].value}, '{sheet_3[row][3].value}', {sheet_3[row][4].value}, {sheet_3[row][5].value}, '{sheet_3[row][6].value}', {sheet_3[row][7].value}, '{sheet_3[row][8].value}', {sheet_3[row][9].value})""")
    DB.connection.commit()

def InsertProducts():
    for row in range(1, sheet_4.max_row):
        DB.cur.execute(
            f"INSERT INTO public.\"Products\" (\"Title\", \"ProductTypeId\", \"ArticleNumber\", \"Description\", \"Image\", \"ProductionPersonCount\", \"ProductionWorkshopNumber\", \"MinCostForAgent\") VALUES ('{sheet_4[row][1].value}', {sheet_4[row][2].value}, '{sheet_4[row][3].value}', '{sheet_4[row][4].value}', '{sheet_4[row][5].value}', {sheet_4[row][6].value}, {sheet_4[row][7].value}, {sheet_4[row][8].value})")
    DB.connection.commit()

def InsertProductMaterials():
    for row in range(1, sheet_5.max_row):
        DB.cur.execute(
            f"INSERT INTO public.\"ProductMaterials\" (\"ProductId\", \"MaterialId\", \"Count\") VALUES ({sheet_5[row][0].value}, {sheet_5[row][1].value}, {sheet_5[row][0].value})")
    DB.connection.commit()